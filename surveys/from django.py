from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count, Avg, Max
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.core.mail import send_mail
from django.conf import settings
from django.utils import timezone
from datetime import datetime, timedelta
import json
from .models import Survey, Question, SurveyResponse, Answer, QuestionChoice, SurveyTemplate, SurveyInvitation, SurveySendLog
from .forms import SurveyForm, QuestionForm, ChoiceForm, SurveyTemplateForm
from graduates.models import Graduate
from .whatsapp_service import SurveySender
import secrets

@login_required
def surveys_home(request):
    """صفحة إدارة الاستبيانات الرئيسية"""
    total_surveys = Survey.objects.count()
    active_surveys = Survey.objects.filter(status='active').count()
    total_responses = SurveyResponse.objects.count()
    sent_surveys = Survey.objects.filter(responses__isnull=False).distinct().count()
    pending_surveys = Survey.objects.filter(status='active', responses__isnull=True).distinct().count()
    
    # حساب معدل الاستجابة
    total_sent = SurveyInvitation.objects.filter(status='sent').count()
    response_rate = round((total_responses / total_sent * 100) if total_sent > 0 else 0, 1)
    
    context = {
        'total_surveys': total_surveys,
        'active_surveys': active_surveys,
        'total_responses': total_responses,
        'sent_surveys': sent_surveys,
        'pending_surveys': pending_surveys,
        'response_rate': response_rate,
    }
    return render(request, 'surveys/surveys_home.html', context)

@login_required
def survey_list(request):
    """قائمة الاستبيانات"""
    surveys = Survey.objects.all().order_by('-created_at')
    
    # البحث
    search_query = request.GET.get('search')
    if search_query:
        surveys = surveys.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # الفلترة حسب الحالة
    status_filter = request.GET.get('status')
    if status_filter == 'active':
        surveys = surveys.filter(status='active')
    elif status_filter == 'inactive':
        surveys = surveys.filter(status='closed')
    
    # التقسيم إلى صفحات
    paginator = Paginator(surveys, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
    }
    return render(request, 'surveys/survey_list.html', context)

@login_required
def survey_create(request):
    """إنشاء استبيان جديد"""
    if request.method == 'POST':
        form = SurveyForm(request.POST)
        if form.is_valid():
            survey = form.save(commit=False)
            survey.created_by = request.user
            survey.save()
            
            # حفظ الخريجين المختارين
            selected_graduates = form.cleaned_data.get('selected_graduates', [])
            
            # إرسال تلقائي إذا كان مفعلاً
            if form.cleaned_data.get('auto_send') and selected_graduates:
                from .whatsapp_service import SurveySender
                sender = SurveySender()
                results = sender.send_survey(survey, selected_graduates, request)
                
                # إظهار رسائل النجاح والفشل
                success_count = results['total_sent']
                failed_count = results['total_failed']
                
                if success_count > 0:
                    messages.success(request, f'تم إنشاء الاستبيان وإرساله إلى {success_count} خريج بنجاح!')
                
                if failed_count > 0:
                    messages.warning(request, f'فشل في إرسال الاستبيان إلى {failed_count} خريج.')
                
                # إظهار تفاصيل الإرسال
                if results['email']:
                    email_success = sum(1 for r in results['email'] if r['success'])
                    email_failed = len(results['email']) - email_success
                    if email_success > 0:
                        messages.info(request, f'تم إرسال {email_success} بريد إلكتروني بنجاح.')
                    if email_failed > 0:
                        messages.warning(request, f'فشل في إرسال {email_failed} بريد إلكتروني.')
                
                if results['whatsapp']:
                    whatsapp_success = sum(1 for r in results['whatsapp'] if r['success'])
                    whatsapp_failed = len(results['whatsapp']) - whatsapp_success
                    if whatsapp_success > 0:
                        messages.info(request, f'تم إنشاء {whatsapp_success} رابط واتساب.')
                    if whatsapp_failed > 0:
                        messages.warning(request, f'فشل في إنشاء {whatsapp_failed} رابط واتساب.')
            else:
                messages.success(request, 'تم إنشاء الاستبيان بنجاح!')
            
            return redirect('surveys:detail', pk=survey.pk)
    else:
        form = SurveyForm()
    
    return render(request, 'surveys/survey_form.html', {'form': form, 'title': 'إنشاء استبيان جديد'})

@login_required
def survey_detail(request, pk):
    """تفاصيل الاستبيان"""
    survey = get_object_or_404(Survey, pk=pk)
    questions = survey.questions.all().order_by('order')
    responses_count = survey.responses.count()
    
    context = {
        'survey': survey,
        'questions': questions,
        'responses_count': responses_count,
    }
    return render(request, 'surveys/survey_detail.html', context)

@login_required
def survey_update(request, pk):
    """تعديل استبيان"""
    survey = get_object_or_404(Survey, pk=pk)
    
    if request.method == 'POST':
        form = SurveyForm(request.POST, instance=survey)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تحديث الاستبيان بنجاح!')
            return redirect('surveys:detail', pk=survey.pk)
    else:
        form = SurveyForm(instance=survey)
    
    return render(request, 'surveys/survey_form.html', {'form': form, 'title': 'تعديل الاستبيان'})

@login_required
def survey_delete(request, pk):
    """حذف استبيان"""
    survey = get_object_or_404(Survey, pk=pk)
    
    if request.method == 'POST':
        survey.delete()
        messages.success(request, 'تم حذف الاستبيان بنجاح!')
        return redirect('surveys:list')
    
    return render(request, 'surveys/survey_confirm_delete.html', {'survey': survey})

@login_required
def question_create(request, survey_pk):
    """إضافة سؤال جديد للاستبيان"""
    survey = get_object_or_404(Survey, pk=survey_pk)
    
    if request.method == 'POST':
        form = QuestionForm(request.POST)
        if form.is_valid():
            question = form.save(commit=False)
            question.survey = survey
            # تحديد ترتيب السؤال
            last_order = survey.questions.aggregate(max_order=Max('order'))['max_order'] or 0
            question.order = last_order + 1
            question.save()
            messages.success(request, 'تم إضافة السؤال بنجاح!')
            return redirect('surveys:detail', pk=survey.pk)
    else:
        form = QuestionForm()
    
    context = {
        'form': form,
        'survey': survey,
        'title': 'إضافة سؤال جديد'
    }
    return render(request, 'surveys/question_form.html', context)

def send_survey(request, pk):
    """إرسال الاستبيان إلى الخريجين المستهدفين عبر البريد الإلكتروني"""
    survey = get_object_or_404(Survey, pk=pk)
    
    # تحديد الخريجين المستهدفين (حسب الكلية أو غير ذلك)
    if hasattr(survey, 'target_college') and survey.target_college:  # تحقق من وجود الحقل والقيمة
        graduates = Graduate.objects.filter(is_active=True, college__iexact=survey.target_college).exclude(email__isnull=True).exclude(email='')
        if not graduates.exists():
            messages.warning(request, f'لا يوجد خريجون في الكلية المستهدفة: {survey.target_college}')
            return redirect('surveys:detail', pk=survey.pk)
    else:
        graduates = Graduate.objects.filter(is_active=True).exclude(email__isnull=True).exclude(email='')  # جميع الخريجين النشطين
    
    # معالجة الإرسال المجمع
    successful_sends = 0
    failed_sends = []
    
    for graduate in graduates:
        # توليد رمز دعوة فريد
        token = secrets.token_urlsafe(16)
        invitation = SurveyInvitation.objects.create(
            survey=survey,
            graduate=graduate,
            invitation_token=token,
            status='sent'  # تعيين الحالة إلى "تم الإرسال" مباشرةً
        )
        survey_url = request.build_absolute_uri(f'/surveys/{survey.pk}/take-public/?token={token}')
        
        try:
            # إرسال البريد الإلكتروني
            subject = survey.get_email_subject()
            message = survey.get_email_message(graduate).replace('http://127.0.0.1:8000/surveys/{}/take-public/'.format(survey.pk), survey_url)  # Replace placeholder
            send_mail(subject, message, None, [graduate.email], fail_silently=False)
            
            successful_sends += 1
            # تسجيل الإرسال الناجح
            SurveySendLog.objects.create(survey=survey, graduate=graduate, send_method='email', status='sent')
            
            # تحديث حالة الدعوة
            invitation.sent_at = timezone.now()
            invitation.save()
            
        except Exception as e:
            failed_sends.append((graduate, str(e)))
            # تسجيل الإرسال الفاشل
            SurveySendLog.objects.create(survey=survey, graduate=graduate, send_method='email', status='failed', error_message=str(e))
        
    # تحديث عدد الرسائل المرسلة في الاستبيان
    survey.total_sent = successful_sends
    survey.email_sent = successful_sends
    survey.save()
    
    if successful_sends > 0:
        messages.success(request, f'تم إرسال الاستبيان إلى {successful_sends} خريج بنجاح!')
    if failed_sends:
        messages.error(request, f'فشل إرسال الاستبيان إلى {len(failed_sends)} خريج. التفاصيل: {", ".join([f"{g.full_name}: {e}" for g, e in failed_sends])}')
    return redirect('surveys:detail', pk=survey.pk)

@login_required
def take_survey(request, pk):
    """عرض الاستبيان للمستلمين للإجابة عليه"""
    survey = get_object_or_404(Survey, pk=pk)
    
    # التحقق من أن الاستبيان نشط
    if survey.status != 'active':
        messages.error(request, 'هذا الاستبيان غير متاح حالياً')
        return redirect('surveys:surveys_home')
    
    # التحقق من تاريخ انتهاء الاستبيان
    if survey.end_date < timezone.now():
        messages.error(request, 'انتهت فترة هذا الاستبيان')
        return redirect('surveys:surveys_home')
    
    # التحقق من وجود إجابة سابقة
    existing_response = None
    if request.user.is_authenticated:
        try:
            graduate = Graduate.objects.get(user=request.user)
            existing_response = SurveyResponse.objects.filter(survey=survey, graduate=graduate).first()
        except Graduate.DoesNotExist:
            pass
    
    if request.method == 'POST':
        # معالجة الإجابة
        if existing_response:
            messages.error(request, 'لقد أجبت على هذا الاستبيان مسبقاً')
            return redirect('surveys:take_survey', pk=pk)
        
        try:
            # إنشاء إجابة جديدة
            graduate = Graduate.objects.get(user=request.user)
            response = SurveyResponse.objects.create(
                survey=survey,
                graduate=graduate,
                is_complete=True
            )
            
            # معالجة إجابات الأسئلة
            for question in survey.questions.all():
                answer_text = request.POST.get(f'question_{question.id}')
                Answer.objects.create(
                    response=response,
                    question=question,
                    answer_text=answer_text or ''
                )

            # تحديث إحصائيات الاستبيان
            survey.responses_received += 1
            survey.save()
            
            messages.success(request, 'تم إرسال إجابتك بنجاح! شكراً لك على المشاركة.')
            return redirect('surveys:survey_thank_you', pk=pk)
            

        except Graduate.DoesNotExist:
            messages.error(request, 'يجب أن تكون خريج للإجابة على الاستبيان')
        except Exception as e:
            messages.error(request, f'حدث خطأ أثناء حفظ الإجابة: {str(e)}')
    
    context = {
        'survey': survey,
        'existing_response': existing_response,
    }
    
    return render(request, 'surveys/take_survey.html', context)

@login_required
def survey_thank_you(request, pk):
    """صفحة شكر بعد الإجابة على الاستبيان"""
    survey = get_object_or_404(Survey, pk=pk)
    
    context = {
        'survey': survey,
    }
    
    return render(request, 'surveys/survey_thank_you.html', context)

@login_required
def survey_results(request, pk):
    """نتائج الاستبيان"""
    survey = get_object_or_404(Survey, pk=pk)
    questions = survey.questions.all().order_by('order')
    surveyresponses = survey.surveyresponses.all()
    
    # تحليل النتائج لكل سؤال
    results = []
    for question in questions:
        answers = Answer.objects.filter(question=question)
        
        if question.question_type in ['multiple_choice', 'single_choice']:
            # تحليل الخيارات
            choices = question.choices.split('\n') if question.choices else []
            choice_counts = {}
            for choice in choices:
                choice_counts[choice] = answers.filter(answer_text=choice).count()
            
            results.append({
                'question': question,
                'type': 'choices',
                'data': choice_counts,
                'total_answers': answers.count()
            })
        else:
            # الأسئلة النصية
            text_answers = list(answers.values_list('answer_text', flat=True))
            results.append({
                'question': question,
                'type': 'text',
                'answers': text_answers,
                'total_answers': len(text_answers)
            })
    
    context = {
        'survey': survey,
        'results': results,
        'total_responses': surveyresponses.count(),
    }
    return render(request, 'surveys/survey_results.html', context)

@login_required
def survey_analytics(request):
    """تحليلات الاستبيانات"""
    # إحصائيات عامة
    total_surveys = Survey.objects.count()
    total_responses = SurveyResponse.objects.count()
    avg_responses = round(total_responses / total_surveys if total_surveys > 0 else 0, 1)
    
    # أكثر الاستبيانات استجابة
    top_surveys = Survey.objects.annotate(
        response_count=Count('surveyresponses')
    ).order_by('-response_count')[:5]
    
    # إحصائيات شهرية
    monthly_stats = SurveyResponse.objects.extra(
        select={'month': 'MONTH(created_at)', 'year': 'YEAR(created_at)'}
    ).values('month', 'year').annotate(count=Count('id')).order_by('-year', '-month')[:12]
    
    context = {
        'total_surveys': total_surveys,
        'total_responses': total_responses,
        'avg_responses': avg_responses,
        'top_surveys': top_surveys,
        'monthly_stats': monthly_stats,
    }
    return render(request, 'surveys/survey_analytics.html', context)

@login_required
def survey_templates(request):
    """قوالب الاستبيانات الجاهزة"""
    templates = [
        {
            'title': 'استبيان رضا الخريجين',
            'description': 'قالب لقياس مستوى رضا الخريجين عن البرنامج الأكاديمي',
            'questions': [
                'ما مدى رضاك عن جودة التعليم في البرنامج؟',
                'هل تشعر أن البرنامج أعدك جيداً لسوق العمل؟',
                'ما تقييمك لأعضاء هيئة التدريس؟',
                'هل تنصح الآخرين بالالتحاق بهذا البرنامج؟'
            ]
        },
        {
            'title': 'استبيان متابعة التوظيف',
            'description': 'قالب لمتابعة الوضع الوظيفي للخريجين',
            'questions': [
                'ما هو وضعك الوظيفي الحالي؟',
                'في أي مجال تعمل حالياً؟',
                'ما مدى ارتباط عملك بتخصصك الأكاديمي؟',
                'كم استغرق الوقت للحصول على أول وظيفة؟'
            ]
        },
        {
            'title': 'استبيان تقييم المناهج',
            'description': 'قالب لتقييم المناهج الدراسية من وجهة نظر الخريجين',
            'questions': [
                'ما تقييمك لحداثة المناهج الدراسية؟',
                'هل كانت المناهج تواكب متطلبات سوق العمل؟',
                'ما المواد التي تشعر أنها كانت الأكثر فائدة؟',
                'ما المواد التي تقترح إضافتها للبرنامج؟'
            ]
        }
    ]
    
    context = {'templates': templates}
    return render(request, 'surveys/survey_templates.html', context)

@login_required
@require_http_methods(["GET"])
def api_survey_chart_data(request, pk):
    """API لبيانات الرسم البياني للاستبيان"""
    survey = get_object_or_404(Survey, pk=pk)
    question_id = request.GET.get('question_id')
    
    if question_id:
        question = get_object_or_404(Question, pk=question_id, survey=survey)
        answers = Answer.objects.filter(question=question)
        
        if question.question_type in ['multiple_choice', 'single_choice']:
            choices = question.choices.split('\n') if question.choices else []
            data = {
                'labels': choices,
                'data': [answers.filter(answer_text=choice).count() for choice in choices]
            }
        else:
            data = {'error': 'نوع السؤال لا يدعم الرسوم البيانية'}
    else:
        data = {'error': 'معرف السؤال مطلوب'}
    
    return JsonResponse(data)



# ========== Views للتحليلات التفصيلية ==========

@login_required
def survey_analytics_detail(request, pk):
    """تحليلات تفصيلية لاستبيان معين"""
    survey = get_object_or_404(Survey, pk=pk)
    surveyresponses = survey.surveyresponses.all()
    
    # إحصائيات أساسية
    total_responses = surveyresponses.count()
    completed_responses = surveyresponses.filter(is_complete=True).count()
    partial_responses = total_responses - completed_responses
    completion_percentage = (completed_responses / total_responses * 100) if total_responses > 0 else 0
    
    # متوسط الوقت
    completed_responses_with_time = surveyresponses.filter(
        is_complete=True, 
        started_at__isnull=False, 
        submitted_at__isnull=False
    )
    
    average_completion_time = 0
    if completed_responses_with_time.exists():
        total_time = sum([
            (response.submitted_at - response.started_at).total_seconds() / 60
            for response in completed_responses_with_time
        ])
        average_completion_time = round(total_time / completed_responses_with_time.count(), 1)
    
    # تحليل الأسئلة
    questions_analysis = []
    for question in survey.questions.all().order_by('order'):
        answers = Answer.objects.filter(question=question)
        total_question_responses = answers.count()
        response_rate = (total_question_responses / total_responses * 100) if total_responses > 0 else 0
        
        analysis = {
            'question_id': question.id,
            'question_text': question.question_text,
            'question_type': question.question_type,
            'total_responses': total_question_responses,
            'response_rate': response_rate,
        }
        
        if question.question_type in ['radio', 'checkbox']:
            # تحليل الخيارات
            choices_data = []
            for choice in question.choices.all():
                choice_count = answers.filter(selected_choices__contains=str(choice.id)).count()
                choices_data.append({
                    'label': choice.choice_text,
                    'count': choice_count
                })
            analysis['choices_data'] = choices_data
            
            # الإجابة الأكثر شيوعاً
            if choices_data:
                most_common = max(choices_data, key=lambda x: x['count'])
                analysis['most_common_answer'] = most_common['label']
        
        elif question.question_type == 'rating':
            # تحليل التقييمات
            ratings = [int(answer.answer_text) for answer in answers if answer.answer_text.isdigit()]
            if ratings:
                analysis['average_rating'] = sum(ratings) / len(ratings)
        
        elif question.question_type in ['text', 'textarea']:
            # عينة من الإجابات النصية
            text_answers = answers.values_list('answer_text', flat=True)[:5]
            analysis['text_responses_count'] = total_question_responses
            analysis['text_samples'] = list(text_answers)
        
        questions_analysis.append(analysis)
    
    # بيانات الخط الزمني للردود
    timeline_data = []
    timeline_labels = []
    
    # آخر 30 يوم
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=29)
    
    current_date = start_date
    while current_date <= end_date:
        day_responses = surveyresponses.filter(
            submitted_at__date=current_date
        ).count()
        timeline_data.append(day_responses)
        timeline_labels.append(current_date.strftime('%m-%d'))
        current_date += timedelta(days=1)
    
    # أحدث الردود
    recent_responses = surveyresponses.order_by('-submitted_at')[:10]
    
    context = {
        'survey': survey,
        'total_responses': total_responses,
        'completed_responses': completed_responses,
        'partial_responses': partial_responses,
        'completion_percentage': completion_percentage,
        'average_completion_time': average_completion_time,
        'questions_analysis': questions_analysis,
        'timeline_data': timeline_data,
        'timeline_labels': timeline_labels,
        'recent_responses': recent_responses,
    }
    return render(request, 'surveys/survey_analytics_detail.html', context)

# ========== Views لإدارة القوالب ==========

@login_required
def survey_templates_list(request):
    """قائمة قوالب الاستبيانات"""
    # القوالب المخصصة
    custom_templates = SurveyTemplate.objects.filter(
        created_by=request.user
    ).order_by('-created_at')
    
    context = {
        'custom_templates': custom_templates,
    }
    return render(request, 'surveys/survey_templates.html', context)

@login_required
def survey_template_create(request):
    """إنشاء قالب استبيان جديد"""
    if request.method == 'POST':
        # معالجة البيانات المرسلة عبر AJAX
        data = json.loads(request.body)
        
        template = SurveyTemplate.objects.create(
            title=data['title'],
            description=data['description'],
            category=data.get('category', 'general'),
            difficulty=data.get('difficulty', 'medium'),
            is_public=data.get('is_public', False),
            created_by=request.user
        )
        
        # إضافة الأسئلة
        for question_data in data['questions']:
            question = Question.objects.create(
                template=template,
                question_text=question_data['question_text'],
                question_type=question_data['question_type'],
                help_text=question_data.get('help_text', ''),
                is_required=question_data.get('is_required', False),
                order=question_data['order']
            )
            
            # إضافة الخيارات إذا كانت موجودة
            for choice_data in question_data.get('choices', []):
                QuestionChoice.objects.create(
                    question=question,
                    choice_text=choice_data['choice_text'],
                    order=choice_data['order']
                )
        
        return JsonResponse({'success': True})
    
    return render(request, 'surveys/survey_template_form.html')

@login_required
def survey_template_edit(request, pk):
    """تعديل قالب استبيان"""
    template = get_object_or_404(SurveyTemplate, pk=pk, created_by=request.user)
    
    if request.method == 'POST':
        # معالجة التعديل
        data = json.loads(request.body)
        
        template.title = data['title']
        template.description = data['description']
        template.category = data.get('category', 'general')
        template.difficulty = data.get('difficulty', 'medium')
        template.is_public = data.get('is_public', False)
        template.save()
        
        # حذف الأسئلة القديمة وإضافة الجديدة
        template.questions.all().delete()
        
        for question_data in data['questions']:
            question = Question.objects.create(
                template=template,
                question_text=question_data['question_text'],
                question_type=question_data['question_type'],
                help_text=question_data.get('help_text', ''),
                is_required=question_data.get('is_required', False),
                order=question_data['order']
            )
            
            for choice_data in question_data.get('choices', []):
                QuestionChoice.objects.create(
                    question=question,
                    choice_text=choice_data['choice_text'],
                    order=choice_data['order']
                )
        
        return JsonResponse({'success': True})
    
    context = {'template': template}
    return render(request, 'surveys/survey_template_form.html', context)

@login_required
def survey_template_delete(request):
    """حذف قالب استبيان"""
    if request.method == 'POST':
        template_id = request.POST.get('template_id')
        template = get_object_or_404(SurveyTemplate, pk=template_id, created_by=request.user)
        template.delete()
        return JsonResponse({'success': True})
    
    return JsonResponse({'success': False})

@login_required
def survey_template_preview(request):
    """معاينة قالب استبيان"""
    template_type = request.GET.get('template')
    
    # قوالب جاهزة
    predefined_templates = {
        'graduate_evaluation': {
            'title': 'استبيان تقييم الخريجين',
            'questions': [
                'ما مدى رضاك عن جودة التعليم في البرنامج؟',
                'هل تشعر أن البرنامج أعدك جيداً لسوق العمل؟',
                'ما تقييمك لأعضاء هيئة التدريس؟',
                'هل تنصح الآخرين بالالتحاق بهذا البرنامج؟'
            ]
        },
        'employment_follow_up': {
            'title': 'استبيان متابعة التوظيف',
            'questions': [
                'ما هو وضعك الوظيفي الحالي؟',
                'في أي مجال تعمل حالياً؟',
                'ما مدى ارتباط عملك بتخصصك الأكاديمي؟',
                'كم استغرق الوقت للحصول على أول وظيفة؟'
            ]
        }
    }
    
    template_data = predefined_templates.get(template_type, {})
    return render(request, 'surveys/template_preview.html', {'template': template_data})

# ========== Views لإدارة الأسئلة والخيارات ==========

@login_required
def question_edit(request, survey_pk, question_pk):
    """تعديل سؤال"""
    survey = get_object_or_404(Survey, pk=survey_pk)
    question = get_object_or_404(Question, pk=question_pk, survey=survey)
    
    if request.method == 'POST':
        form = QuestionForm(request.POST, instance=question)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تحديث السؤال بنجاح!')
            return redirect('surveys:detail', pk=survey.pk)
    else:
        form = QuestionForm(instance=question)
    
    context = {
        'form': form,
        'survey': survey,
        'question': question,
    }
    return render(request, 'surveys/question_form.html', context)

@login_required
def question_delete(request, survey_pk, question_pk):
    """حذف سؤال"""
    survey = get_object_or_404(Survey, pk=survey_pk)
    question = get_object_or_404(Question, pk=question_pk, survey=survey)
    
    # إحصائيات للعرض
    total_answers = Answer.objects.filter(question=question).count()
    unique_respondents = Answer.objects.filter(question=question).values('surveyresponse__graduate').distinct().count()
    total_survey_responses = survey.surveyresponses.count()
    response_rate = (unique_respondents / total_survey_responses * 100) if total_survey_responses > 0 else 0
    
    if request.method == 'POST':
        question.delete()
        messages.success(request, 'تم حذف السؤال بنجاح!')
        return redirect('surveys:detail', pk=survey.pk)
    
    context = {
        'survey': survey,
        'question': question,
        'total_answers': total_answers,
        'unique_respondents': unique_respondents,
        'response_rate': response_rate,
    }
    return render(request, 'surveys/question_confirm_delete.html', context)

@login_required
def choice_create(request, survey_pk, question_pk):
    """إضافة خيار جديد لسؤال"""
    survey = get_object_or_404(Survey, pk=survey_pk)
    question = get_object_or_404(Question, pk=question_pk, survey=survey)
    
    # تحديد الترتيب التالي
    next_order = (question.choices.aggregate(Max('order'))['order__max'] or 0) + 1
    
    if request.method == 'POST':
        form = ChoiceForm(request.POST)
        if form.is_valid():
            choice = form.save(commit=False)
            choice.question = question
            choice.save()
            messages.success(request, 'تم إضافة الخيار بنجاح!')
            return redirect('surveys:question_edit', survey_pk=survey.pk, question_pk=question.pk)
    else:
        form = ChoiceForm(initial={'order': next_order})
    
    context = {
        'form': form,
        'question': question,
        'next_order': next_order,
    }
    return render(request, 'surveys/choice_form.html', context)

@login_required
def choice_edit(request, survey_pk, question_pk, choice_pk):
    """تعديل خيار"""
    survey = get_object_or_404(Survey, pk=survey_pk)
    question = get_object_or_404(Question, pk=question_pk, survey=survey)
    choice = get_object_or_404(QuestionChoice, pk=choice_pk, question=question)
    
    if request.method == 'POST':
        form = ChoiceForm(request.POST, instance=choice)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تحديث الخيار بنجاح!')
            return redirect('surveys:question_edit', survey_pk=survey.pk, question_pk=question.pk)
    else:
        form = ChoiceForm(instance=choice)
    
    context = {
        'form': form,
        'question': question,
        'choice': choice,
    }
    return render(request, 'surveys/choice_form.html', context)

@login_required
def choice_delete(request, survey_pk, question_pk, choice_pk):
    """حذف خيار"""
    survey = get_object_or_404(Survey, pk=survey_pk)
    question = get_object_or_404(Question, pk=question_pk, survey=survey)
    choice = get_object_or_404(QuestionChoice, pk=choice_pk, question=question)
    
    # إحصائيات الاستخدام
    choice_selections = Answer.objects.filter(
        question=question,
        selected_choices__contains=str(choice.id)
    ).count()
    
    unique_selectors = Answer.objects.filter(
        question=question,
        selected_choices__contains=str(choice.id)
    ).values('surveyresponse__graduate').distinct().count()
    
    total_question_responses = Answer.objects.filter(question=question).count()
    selection_percentage = (choice_selections / total_question_responses * 100) if total_question_responses > 0 else 0
    
    # الخيارات المتبقية
    remaining_choices = question.choices.exclude(pk=choice.pk)
    
    if request.method == 'POST':
        choice.delete()
        messages.success(request, 'تم حذف الخيار بنجاح!')
        return redirect('surveys:question_edit', survey_pk=survey.pk, question_pk=question.pk)
    
    context = {
        'choice': choice,
        'choice_selections': choice_selections,
        'unique_selectors': unique_selectors,
        'selection_percentage': selection_percentage,
        'remaining_choices': remaining_choices,
    }
    return render(request, 'surveys/choice_confirm_delete.html', context)

# ========== Views لإدارة الردود ==========

@login_required
def survey_response_list(request, survey_pk):
    """قائمة ردود استبيان معين"""
    survey = get_object_or_404(Survey, pk=survey_pk)
    surveyresponses = survey.surveyresponses.all()
    
    # إحصائيات أساسية
    total_responses = surveyresponses.count()
    completed_responses = surveyresponses.filter(is_complete=True).count()
    partial_responses = total_responses - completed_responses
    completion_percentage = (completed_responses / total_responses * 100) if total_responses > 0 else 0
    
    # متوسط الوقت
    average_time = 0
    completed_with_time = surveyresponses.filter(
        is_complete=True,
        started_at__isnull=False,
        submitted_at__isnull=False
    )
    if completed_with_time.exists():
        total_time = sum([
            (r.submitted_at - r.started_at).total_seconds() / 60
            for r in completed_with_time
        ])
        average_time = round(total_time / completed_with_time.count(), 1)
    
    # فلترة وبحث
    search_query = request.GET.get('search')
    if search_query:
        surveyresponses = surveyresponses.filter(
            Q(graduate__full_name__icontains=search_query) |
            Q(graduate__email__icontains=search_query)
        )
    
    status_filter = request.GET.get('status')
    if status_filter == 'completed':
        surveyresponses = surveyresponses.filter(is_complete=True)
    elif status_filter == 'partial':
        surveyresponses = surveyresponses.filter(is_complete=False, started_at__isnull=False)
    elif status_filter == 'not_started':
        surveyresponses = surveyresponses.filter(started_at__isnull=True)
    
    # فلترة بالتاريخ
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        surveyresponses = surveyresponses.filter(submitted_at__date__gte=date_from)
    if date_to:
        surveyresponses = surveyresponses.filter(submitted_at__date__lte=date_to)
    
    # ترتيب
    sort_by = request.GET.get('sort_by', '-submitted_at')
    surveyresponses = surveyresponses.order_by(sort_by)
    
    # تصدير
    export_format = request.GET.get('export')
    if export_format in ['excel', 'pdf']:
        return export_survey_responses(survey, surveyresponses, export_format)
    
    # تقسيم الصفحات
    paginator = Paginator(surveyresponses, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'survey': survey,
        'surveyresponses': page_obj,
        'total_responses': total_responses,
        'completed_responses': completed_responses,
        'partial_responses': partial_responses,
        'completion_percentage': completion_percentage,
        'average_time': average_time,
    }
    return render(request, 'surveys/survey_response_list.html', context)

# ========== Views لدعم Google Forms ==========

@login_required
def create_google_form_survey(request):
    """إنشاء استبيان مرتبط بـ Google Forms"""
    if request.method == 'POST':
        try:
            # التحقق من وجود البيانات المطلوبة
            title = request.POST.get('title')
            description = request.POST.get('description', '')
            google_form_url = request.POST.get('google_form_url')
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            
            if not title or not google_form_url:
                return JsonResponse({
                    'success': False,
                    'error': 'العنوان ورابط جوجل فورم مطلوبان'
                })
            
            # تحويل التواريخ
            try:
                if start_date:
                    start_date = datetime.strptime(start_date, '%Y-%m-%dT%H:%M')
                else:
                    start_date = timezone.now()
                
                if end_date:
                    end_date = datetime.strptime(end_date, '%Y-%m-%dT%H:%M')
                else:
                    end_date = timezone.now() + timedelta(days=30)
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'error': 'صيغة التاريخ غير صحيحة'
                })
            
            survey = Survey.objects.create(
                title=title,
                description=description,
                google_form_url=google_form_url,
                start_date=start_date,
                end_date=end_date,
                created_by=request.user,
                status='active'  # استخدام status بدلاً من is_active
            )
            
            return JsonResponse({
                'success': True,
                'redirect_url': f'/surveys/{survey.id}/'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'حدث خطأ أثناء إنشاء الاستبيان: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'طريقة غير مدعومة'})

# ========== Helper Functions ==========

def export_survey_responses(survey, surveyresponses, format_type):
    """تصدير ردود الاستبيان"""
    if format_type == 'excel':
        # تصدير Excel
        import openpyxl
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"ردود {survey.title}"
        
        # العناوين
        headers = ['الخريج', 'البريد الإلكتروني', 'تاريخ الإرسال', 'الحالة']
        for question in survey.questions.all():
            headers.append(question.question_text[:50])
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # البيانات
        for row, surveyresponse in enumerate(surveyresponses, 2):
            ws.cell(row=row, column=1, value=surveyresponse.graduate.full_name)
            ws.cell(row=row, column=2, value=surveyresponse.graduate.email)
            ws.cell(row=row, column=3, value=surveyresponse.submitted_at.strftime('%Y-%m-%d %H:%M') if surveyresponse.submitted_at else '')
            ws.cell(row=row, column=4, value='مكتمل' if surveyresponse.is_complete else 'جزئي')
            
            # الإجابات
            for col, question in enumerate(survey.questions.all(), 5):
                answer = Answer.objects.filter(surveyresponse=surveyresponse, question=question).first()
                ws.cell(row=row, column=col, value=answer.answer_text if answer else '')
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="survey_responses_{survey.id}.xlsx"'
        wb.save(response)
        return response
    
    elif format_type == 'pdf':
        # تصدير PDF
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter, A4
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="survey_responses_{survey.id}.pdf"'
        
        p = canvas.Canvas(response, pagesize=A4)
        p.drawString(100, 750, f"ردود الاستبيان: {survey.title}")
        
        y = 700
        for resp in surveyresponses[:50]:  # أول 50 رد فقط
            p.drawString(100, y, f"{resp.graduate.full_name} - {resp.graduate.email}")
            y -= 20
            if y < 100:
                p.showPage()
                y = 750
        
        p.save()
        return response

# ========== Views للإجراءات المجمعة ==========

@login_required
def bulk_delete_responses(request):
    """حذف مجمع للردود"""
    if request.method == 'POST':
        response_ids = request.POST.getlist('response_ids')
        deleted_count = SurveyResponse.objects.filter(id__in=response_ids).delete()[0]
        
        return JsonResponse({
            'success': True,
            'deleted_count': deleted_count
        })
    
    return JsonResponse({'success': False})

@login_required
def delete_response(request):
    """حذف رد واحد"""
    if request.method == 'POST':
        response_id = request.POST.get('response_id')
        SurveyResponse.objects.filter(id=response_id).delete()
        return JsonResponse({'success': True})
    
    return JsonResponse({'success': False})

# ========== Views لنظام الإرسال المتقدم ==========

@login_required
def get_graduates(request):
    """الحصول على الخريجين حسب الفلتر"""
    if request.method == 'POST':
        try:
            # استخراج الفلاتر
            college = request.POST.get('college', '')
            major = request.POST.get('major', '')
            year = request.POST.get('year', '')
            status = request.POST.get('status', '')
            
            # بناء الاستعلام
            graduates = Graduate.objects.filter(is_active=True)
            
            if college:
                graduates = graduates.filter(college=college)
            if major:
                graduates = graduates.filter(major=major)
            if year:
                graduates = graduates.filter(graduation_year=year)
            if status:
                graduates = graduates.filter(employment_status=status)
            
            # تحويل البيانات
            graduates_data = []
            for graduate in graduates:
                graduates_data.append({
                    'id': graduate.id,
                    'full_name': graduate.full_name,
                    'email': graduate.email,
                    'phone': graduate.phone,
                    'college': graduate.college or '',
                    'major': graduate.major or '',
                    'graduation_year': graduate.graduation_year or '',
                    'employment_status': graduate.get_employment_status_display() or ''
                })
            
            return JsonResponse({
                'success': True,
                'graduates': graduates_data,
                'count': len(graduates_data)
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'طريقة غير مدعومة'})


@login_required
def send_survey_bulk(request):
    """إرسال استبيان لعدة خريجين مع توليد دعوة فريدة لكل خريج"""
    if request.method == 'POST':
        try:
            survey_id = request.POST.get('survey_id')
            send_method = request.POST.get('send_method')
            graduate_ids = request.POST.getlist('graduate_ids')
            
            if not survey_id or not send_method or not graduate_ids:
                return JsonResponse({
                    'success': False,
                    'error': 'بيانات غير مكتملة'
                })
            
            # الحصول على الاستبيان
            survey = get_object_or_404(Survey, pk=survey_id)
            
            # الحصول على الخريجين
            graduates = Graduate.objects.filter(id__in=graduate_ids, is_active=True)
            
            # إحصائيات الإرسال
            email_sent = 0
            whatsapp_sent = 0
            failed_sent = 0
            invitations = []
            
            for graduate in graduates:
                try:
                    # توليد رمز دعوة فريد
                    token = secrets.token_urlsafe(16)
                    invitation = SurveyInvitation.objects.create(
                        survey=survey,
                        graduate=graduate,
                        invitation_token=token,
                        status='sent'
                    )
                    invitations.append(invitation)
                    link = request.build_absolute_uri(f'/surveys/{survey.id}/take-public/?token={token}')
                    
                    # إرسال البريد الإلكتروني
                    if send_method in ['email', 'both'] and graduate.email:
                        subject = f'استبيان: {survey.title}'
                        message = f"""
                        عزيزي/عزيزتي {graduate.full_name},\n\nنرجو منك المشاركة في الاستبيان التالي:\n{survey.title}\n\n{survey.description}\n\nللمشاركة، يرجى النقر على الرابط التالي:\n{link}\n\nشكراً لك على وقتك."""
                        send_mail(
                            subject,
                            message,
                            settings.DEFAULT_FROM_EMAIL,
                            [graduate.email],
                            fail_silently=False,
                        )
                        email_sent += 1
                        SurveySendLog.objects.create(
                            survey=survey,
                            graduate=graduate,
                            send_method='email',
                            status='sent'
                        )
                    # إرسال الواتساب (محاكاة)
                    if send_method in ['whatsapp', 'both'] and graduate.phone:
                        # هنا يمكنك ربط خدمة واتساب حقيقية
                        whatsapp_sent += 1
                        SurveySendLog.objects.create(
                            survey=survey,
                            graduate=graduate,
                            send_method='whatsapp',
                            status='sent'
                        )
                except Exception as e:
                    failed_sent += 1
                    SurveySendLog.objects.create(
                        survey=survey,
                        graduate=graduate,
                        send_method=send_method,
                        status='failed',
                        error_message=str(e)
                    )
            
            # تحديث إحصائيات الاستبيان
            survey.total_sent += len(graduates)
            survey.email_sent += email_sent
            survey.whatsapp_sent += whatsapp_sent
            survey.save()
            
            return JsonResponse({
                'success': True,
                'total_sent': len(graduates),
                'email_sent': email_sent,
                'whatsapp_sent': whatsapp_sent,
                'failed_sent': failed_sent,
                'message': f'تم إرسال الاستبيان إلى {len(graduates)} خريج بنجاح'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'خطأ في الإرسال: {str(e)}'
            })
    return JsonResponse({'success': False, 'error': 'طريقة غير مدعومة'})


@login_required
def send_survey_logs(request, survey_id):
    """عرض سجلات إرسال استبيان معين"""
    survey = get_object_or_404(Survey, pk=survey_id)
    logs = survey.send_logs.all().order_by('-sent_at')
    
    # إحصائيات
    total_sent = logs.count()
    successful_sent = logs.filter(status='sent').count()
    failed_sent = logs.filter(status='failed').count()
    
    context = {
        'survey': survey,
        'logs': logs,
        'total_sent': total_sent,
        'successful_sent': successful_sent,
        'failed_sent': failed_sent,
    }
    
    return render(request, 'surveys/send_logs.html', context)


# ========== Views لعرض الاستبيان للمستلمين ==========

def take_survey_public(request, pk):
    """عرض الاستبيان للمستلمين للإجابة عليه (بدون تسجيل دخول) عبر رمز دعوة فريد"""
    survey = get_object_or_404(Survey, pk=pk)
    token = request.GET.get('token')
    invitation = None
    graduate = None
    error_message = None

    # التحقق من وجود التوكن وصحته
    if token:
        try:
            invitation = SurveyInvitation.objects.get(survey=survey, invitation_token=token)
            graduate = invitation.graduate
            if invitation.status == 'completed':
                error_message = 'لقد قمت بالإجابة على هذا الاستبيان مسبقًا.'
        except SurveyInvitation.DoesNotExist:
            error_message = 'رابط الدعوة غير صالح أو منتهي.'
    else:
        error_message = 'رابط الدعوة غير صالح.'

    if error_message:
        return render(request, 'surveys/take_survey_public.html', {
            'survey': survey,
            'error_message': error_message
        })

    if request.method == 'POST':
        try:
            # التحقق من عدم وجود إجابة سابقة
            if invitation.status == 'completed':
                return render(request, 'surveys/take_survey_public.html', {
                    'survey': survey,
                    'error_message': 'لقد قمت بالإجابة على هذا الاستبيان مسبقًا.'
                })
            # إنشاء إجابة جديدة
            response = SurveyResponse.objects.create(
                survey=survey,
                graduate=graduate,
                is_complete=True
            )
            # معالجة إجابات الأسئلة
            for question in survey.questions.all():
                answer_text = request.POST.get(f'question_{question.id}')
                if answer_text:
                    Answer.objects.create(
                        response=response,
                        question=question,
                        answer_text=answer_text
                    )
            # تحديث حالة الدعوة
            invitation.status = 'completed'
            invitation.completed_at = timezone.now()
            invitation.save()
            # تحديث إحصائيات الاستبيان
            survey.responses_received += 1
            survey.save()
            return render(request, 'surveys/survey_thank_you_public.html', {'survey': survey})
        except Exception as e:
            return render(request, 'surveys/take_survey_public.html', {
                'survey': survey,
                'error_message': f'حدث خطأ أثناء حفظ الإجابة: {str(e)}'
            })

    return render(request, 'surveys/take_survey_public.html', {
        'survey': survey,
        'graduate': graduate,
        'invitation': invitation
    })


def survey_thank_you_public(request, pk):
    """صفحة شكر بعد الإجابة على الاستبيان (بدون تسجيل دخول)"""
    survey = get_object_or_404(Survey, pk=pk)
    
    context = {
        'survey': survey,
    }
    
    return render(request, 'surveys/survey_thank_you_public.html', context)

@login_required
def survey_templates_select(request):
    """واجهة اختيار قالب الاستبيان الجاهز أو Google Forms"""
    return render(request, 'surveys/survey_templates_select.html')

@login_required
def survey_create_from_template(request, template_id):
    """إنشاء استبيان جديد من قالب محدد (مثال: رضا عن المناهج) مع إرسال الإيميل للخريجين حسب القسم"""
    from .models import Survey, Question, QuestionChoice, SurveyInvitation, SurveySendLog
    from graduates.models import Graduate
    from django.utils import timezone
    from django.core.mail import send_mail
    import secrets

    # مثال: قالب رضا عن المناهج الدراسية
    if template_id == 'curriculum_satisfaction':
        template = {
            'title': 'استبيان رضا عن المناهج الدراسية',
            'description': 'استبيان لقياس مدى رضا الخريجين عن المناهج الدراسية وجودة البرامج الأكاديمية.',
            'questions': [
                {
                    'text': 'ما مدى رضاك عن المناهج الدراسية في تخصصك؟',
                    'type': 'radio',
                    'required': True,
                    'choices': ['راض جدا', 'راض', 'محايد', 'غير راض', 'غير راض إطلاقاً']
                },
                {
                    'text': 'هل تعتقد أن برامج الجامعة تتوافق مع متطلبات سوق العمل؟',
                    'type': 'radio',
                    'required': True,
                    'choices': ['نعم بشكل كبير', 'نعم إلى حد ما', 'لا، هناك فجوة كبيرة', 'لا أعلم']
                }
            ]
        }
        if request.method == 'POST':
            title = request.POST.get('title') or template['title']
            description = request.POST.get('description') or template['description']
            # تحويل التواريخ من نص إلى كائن تاريخ
            from datetime import datetime
            start_date_str = request.POST.get('start_date')
            end_date_str = request.POST.get('end_date')
            if start_date_str:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            else:
                start_date = timezone.now()
            if end_date_str:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            else:
                end_date = timezone.now() + timezone.timedelta(days=14)
            # خيارات الجمهور المستهدف
            faculty = request.POST.get('faculty')
            major = request.POST.get('major')
            grad_year = request.POST.get('grad_year')
            employment_status = request.POST.getlist('employment_status')
            send_email = request.POST.get('send_email')
            # جلب الخريجين المطابقين
            graduates = Graduate.objects.filter(is_active=True)
            if faculty:
                graduates = graduates.filter(college=faculty)
            if major:
                graduates = graduates.filter(major=major)
            if grad_year:
                graduates = graduates.filter(graduation_year=grad_year)
            if employment_status and 'all' not in employment_status:
                graduates = graduates.filter(employment_status__in=employment_status)
            graduates = graduates.exclude(email__isnull=True).exclude(email='')
            # إنشاء الاستبيان
            survey = Survey.objects.create(
                title=title,
                description=description,
                status='active',
                survey_type='template',
                template_name='رضا عن المناهج الدراسية',
                created_by=request.user,
                start_date=start_date,
                end_date=end_date,
                send_method='email',
                auto_send=True
            )
            for idx, q in enumerate(template['questions'], start=1):
                question = Question.objects.create(
                    survey=survey,
                    question_text=q['text'],
                    question_type=q['type'],
                    is_required=q['required'],
                    order=idx
                )
                for cidx, choice in enumerate(q['choices'], start=1):
                    QuestionChoice.objects.create(
                        question=question,
                        choice_text=choice,
                        order=cidx
                    )
            # إرسال الإيميل للخريجين
            email_sent = 0
            for grad in graduates:
                token = secrets.token_urlsafe(16)
                SurveyInvitation.objects.create(
                    survey=survey,
                    graduate=grad,
                    invitation_token=token,
                    status='sent',
                    sent_at=timezone.now()
                )
                survey_url = request.build_absolute_uri(f'/surveys/{survey.pk}/take-public/?token={token}')
                subject = survey.get_email_subject()
                message = survey.get_email_message(grad).replace('http://127.0.0.1:8000/surveys/{}/take-public/'.format(survey.pk), survey_url)
                try:
                    send_mail(subject, message, None, [grad.email], fail_silently=False)
                    SurveySendLog.objects.create(survey=survey, graduate=grad, send_method='email', status='sent')
                    email_sent += 1
                except Exception as e:
                    SurveySendLog.objects.create(survey=survey, graduate=grad, send_method='email', status='failed', error_message=str(e))
            survey.total_sent = graduates.count()
            survey.email_sent = email_sent
            survey.save()
            messages.success(request, f'تم إنشاء الاستبيان وإرساله إلى {email_sent} خريج بنجاح!')
            return redirect('surveys:detail', pk=survey.pk)
        return render(request, 'surveys/survey_form_from_template.html', {'template': template})
    return redirect('surveys:create')
